"""Decimal-safe CSV/XLSX reader.

Returns rows as ``list[dict[str, str]]`` keyed by the header row, with every
cell normalised to a *string*. We deliberately keep everything as text at this
layer so that downstream parsing into ``Decimal`` is explicit and lossless —
pandas/openpyxl float coercion is the precise thing we are avoiding for
audited financial figures.

For XLSX, a numeric cell that openpyxl has already parsed to ``float`` is
converted via ``Decimal(str(value))`` only as a last resort; integers and
text pass through unchanged. The raw underlying string from the sheet is used
where available.
"""

from __future__ import annotations

import csv
from pathlib import Path


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        # openpyxl handed us a float; render without spurious binary noise.
        # Integral floats become "123", others use repr to preserve digits.
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return [
            {(k or "").strip(): _stringify(v) for k, v in row.items()}
            for row in reader
        ]


def read_xlsx(path: Path) -> list[dict[str, str]]:
    # Imported lazily so CSV-only users don't pay for openpyxl.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [_stringify(c) for c in next(rows_iter)]
        except StopIteration:
            return []
        out: list[dict[str, str]] = []
        for raw in rows_iter:
            cells = [_stringify(c) for c in raw]
            # pad/truncate to header width
            cells += [""] * (len(header) - len(cells))
            if not any(cells):
                continue  # skip fully blank rows
            out.append(dict(zip(header, cells)))
        return out
    finally:
        wb.close()


def read_listing(path: str | Path) -> list[dict[str, str]]:
    """Dispatch on file extension. Raises on unsupported types."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path)
    raise ValueError(f"Unsupported listing type: {suffix!r} ({path.name})")
